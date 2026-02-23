"""Topic and keyword CRUD endpoints."""

import io

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from backend.database import (
    create_topic,
    delete_topic,
    get_all_topics,
    get_db,
    replace_keywords,
    update_topic_name,
)
from backend.models import KeywordsUpdate, TopicCreate, TopicUpdate

router = APIRouter(prefix="/api/topics", tags=["topics"])


@router.get("")
async def list_topics():
    db = await get_db()
    try:
        return await get_all_topics(db)
    finally:
        await db.close()


@router.post("", status_code=201)
async def create_topic_endpoint(body: TopicCreate):
    db = await get_db()
    try:
        return await create_topic(db, body.name, body.keywords)
    except Exception as e:
        if "UNIQUE" in str(e):
            raise HTTPException(400, f"Topic '{body.name}' already exists")
        raise
    finally:
        await db.close()


@router.put("/{topic_id}")
async def update_topic_endpoint(topic_id: int, body: TopicUpdate):
    db = await get_db()
    try:
        if not await update_topic_name(db, topic_id, body.name):
            raise HTTPException(404, "Topic not found")
        return {"ok": True}
    finally:
        await db.close()


@router.delete("/{topic_id}")
async def delete_topic_endpoint(topic_id: int):
    db = await get_db()
    try:
        if not await delete_topic(db, topic_id):
            raise HTTPException(404, "Topic not found")
        return {"ok": True}
    finally:
        await db.close()


@router.get("/export")
async def export_topics_excel():
    db = await get_db()
    try:
        topics = await get_all_topics(db)
    finally:
        await db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Topics"
    ws.append(["Topic", "Keywords"])
    for topic in topics:
        ws.append([topic["name"], ", ".join(topic["keywords"])])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=topics.xlsx"},
    )


@router.post("/import")
async def import_topics_excel(file: UploadFile = File(...)):
    contents = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Invalid Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "File has no data rows")

    db = await get_db()
    try:
        existing = await get_all_topics(db)
        by_name = {t["name"].lower(): t for t in existing}

        created = 0
        updated = 0

        for row in rows[1:]:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            if not name:
                continue
            kw_raw = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            keywords = [k.strip() for k in kw_raw.split(",") if k.strip()]

            if name.lower() in by_name:
                await replace_keywords(db, by_name[name.lower()]["id"], keywords)
                updated += 1
            else:
                await create_topic(db, name, keywords)
                created += 1

        return {"ok": True, "created": created, "updated": updated}
    finally:
        await db.close()


@router.put("/{topic_id}/keywords")
async def update_keywords_endpoint(topic_id: int, body: KeywordsUpdate):
    db = await get_db()
    try:
        if not await replace_keywords(db, topic_id, body.keywords):
            raise HTTPException(404, "Topic not found")
        return {"ok": True}
    finally:
        await db.close()
