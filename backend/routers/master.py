"""Master stakeholder list endpoints."""

import io
import json

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from backend.database import (
    get_db,
    add_to_master_list,
    get_master_list,
    update_master_entry,
    delete_master_entry,
    get_master_result_ids,
    remove_master_activity_by_result,
)

router = APIRouter(prefix="/api/master", tags=["master"])


class MasterAddRequest(BaseModel):
    result_id: int
    member_name: str
    member_id: str | None = None
    party: str = ""
    member_type: str = ""
    constituency: str = ""


class MasterUpdateRequest(BaseModel):
    notes: str | None = None
    priority: str | None = None


@router.post("/add", status_code=201)
async def add_to_master(body: MasterAddRequest):
    db = await get_db()
    try:
        result = await add_to_master_list(
            db,
            member_name=body.member_name,
            member_id=body.member_id,
            party=body.party,
            member_type=body.member_type,
            constituency=body.constituency,
            result_id=body.result_id,
        )
        return result
    finally:
        await db.close()


@router.get("")
async def list_master():
    db = await get_db()
    try:
        return await get_master_list(db)
    finally:
        await db.close()


@router.put("/{master_id}")
async def update_master(master_id: int, body: MasterUpdateRequest):
    db = await get_db()
    try:
        ok = await update_master_entry(db, master_id, notes=body.notes, priority=body.priority)
        if not ok:
            raise HTTPException(404, "Entry not found")
        return {"ok": True}
    finally:
        await db.close()


@router.delete("/{master_id}")
async def remove_from_master(master_id: int):
    db = await get_db()
    try:
        ok = await delete_master_entry(db, master_id)
        if not ok:
            raise HTTPException(404, "Entry not found")
        return {"ok": True}
    finally:
        await db.close()


@router.get("/result-ids")
async def master_result_ids():
    """Get all result IDs linked to master list entries."""
    db = await get_db()
    try:
        ids = await get_master_result_ids(db)
        return {"result_ids": ids}
    finally:
        await db.close()


@router.delete("/activity/{result_id}")
async def remove_activity_by_result(result_id: int):
    """Remove a result link from the master list (granular removal)."""
    db = await get_db()
    try:
        ok = await remove_master_activity_by_result(db, result_id)
        if not ok:
            raise HTTPException(404, "Activity link not found")
        return {"ok": True}
    finally:
        await db.close()


@router.get("/export")
async def export_master():
    """Export master list as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    db = await get_db()
    try:
        entries = await get_master_list(db)
    finally:
        await db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Stakeholder List"

    # Header
    headers = ["Name", "Party", "Type", "Constituency", "Priority", "Notes",
               "Activities", "Topics", "Latest Activity"]
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, entry in enumerate(entries, 2):
        activities = entry.get("activities", [])
        all_topics = set()
        latest_date = ""
        for a in activities:
            try:
                topics = json.loads(a.get("topics", "[]"))
                all_topics.update(topics if isinstance(topics, list) else [topics])
            except (json.JSONDecodeError, TypeError):
                pass
            if a.get("activity_date", "") > latest_date:
                latest_date = a["activity_date"]

        ws.cell(row=row_idx, column=1, value=entry["member_name"])
        ws.cell(row=row_idx, column=2, value=entry.get("party", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("member_type", ""))
        ws.cell(row=row_idx, column=4, value=entry.get("constituency", ""))
        ws.cell(row=row_idx, column=5, value=entry.get("priority", ""))
        ws.cell(row=row_idx, column=6, value=entry.get("notes", ""))
        ws.cell(row=row_idx, column=7, value=len(activities))
        ws.cell(row=row_idx, column=8, value=", ".join(sorted(all_topics)))
        ws.cell(row=row_idx, column=9, value=latest_date)

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="master_stakeholder_list.xlsx"'},
    )
