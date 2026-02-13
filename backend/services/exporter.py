"""Excel export with openpyxl — preserves hyperlinks."""

import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_excel_export(results: list[dict], scan: dict) -> io.BytesIO:
    """Create an Excel workbook from scan results. Returns a BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header styling
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(bottom=Side(style="thin", color="E2E8F0"))

    headers = [
        "Name", "Party", "Type", "Topic(s)",
        "Summary & Date", "Forum", "Quote / Action", "Confidence",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Confidence fills
    conf_fills = {
        "High": PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
        "Medium": PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid"),
        "Low": PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
    }

    for row_idx, r in enumerate(results, 2):
        # Parse topics
        topics = r.get("topics", "")
        try:
            topics = ", ".join(json.loads(topics))
        except (json.JSONDecodeError, TypeError):
            pass

        # Format date
        date_str = r.get("activity_date", "")
        if date_str and len(date_str) >= 10:
            parts = date_str[:10].split("-")
            if len(parts) == 3:
                date_str = f"{parts[2]}/{parts[1]}/{parts[0][2:]}"

        ws.cell(row=row_idx, column=1, value=r.get("member_name", ""))
        ws.cell(row=row_idx, column=2, value=r.get("party", ""))
        ws.cell(row=row_idx, column=3, value=r.get("member_type", ""))
        ws.cell(row=row_idx, column=4, value=topics)
        ws.cell(row=row_idx, column=5, value=f"{r.get('summary', '')} ({date_str})")
        ws.cell(row=row_idx, column=6, value=r.get("forum", ""))

        # Quote with hyperlink
        quote_cell = ws.cell(row=row_idx, column=7)
        quote_text = r.get("verbatim_quote", "") or ""
        quote_cell.value = quote_text
        source_url = r.get("source_url", "")
        if source_url:
            quote_cell.hyperlink = source_url
            quote_cell.font = Font(color="0563C1", underline="single")

        # Confidence with color
        conf = r.get("confidence", "")
        conf_cell = ws.cell(row=row_idx, column=8, value=conf)
        if conf in conf_fills:
            conf_cell.fill = conf_fills[conf]
        conf_cell.alignment = Alignment(horizontal="center")

        # Border on all cells
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Column widths
    col_widths = [22, 18, 8, 22, 50, 25, 60, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A1:H{len(results) + 1}"

    # Freeze header
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
