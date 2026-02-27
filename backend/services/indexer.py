"""Member Topic Index — scoring engine and Excel export."""

import json
from collections import defaultdict
from io import BytesIO


CONFIDENCE_WEIGHTS = {"High": 3, "Medium": 2, "Low": 1}

SOURCE_LABELS = {
    "hansard": "Hansard",
    "written_question": "Written Q",
    "written_statement": "Written Stmt",
    "edm": "EDM",
    "bill": "Bill",
    "division": "Division",
}


def _activity_record(r: dict) -> dict:
    """Extract the fields needed for the frontend activity expansion."""
    return {
        "activity_date": r.get("activity_date") or "",
        "source_type": r.get("source_type") or "",
        "forum": r.get("forum") or "",
        "summary": r.get("summary") or "",
        "verbatim_quote": r.get("verbatim_quote") or "",
        "source_url": r.get("source_url") or "",
        "confidence": r.get("confidence") or "",
    }


def generate_index(results: list[dict]) -> dict:
    """Aggregate results into a ranked member-by-topic index.

    Deduplicates by dedup_key across scans so overlapping scans don't
    double-count the same parliamentary contribution.

    Returns:
        {
            "topics": {topic_name: [ranked_member_dict, ...]},
            "cross_topic": [ranked_member_dict, ...],
            "meta": {"total_results": int, "total_members": int, "topics": [str]}
        }
    Each member dict includes an "activities" list for the expansion panel.
    """
    # Deduplicate by dedup_key — keep most-recently-seen per key
    seen: dict[str, dict] = {}
    for r in results:
        key = r.get("dedup_key") or str(r.get("id"))
        seen[key] = r
    deduped = list(seen.values())

    # Per-topic accumulators: {topic: {member_name: {...}}}
    topic_members: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(lambda: {
            "score": 0,
            "mentions": 0,
            "sources": defaultdict(int),
            "latest_date": "",
            "member_name": "",
            "party": "",
            "member_type": "",
            "activities": [],
        })
    )

    # Cross-topic accumulators: {member_name: {...}}
    cross_members: dict[str, dict] = defaultdict(lambda: {
        "score": 0,
        "mentions": 0,
        "topics_active": set(),
        "member_name": "",
        "party": "",
        "member_type": "",
        "activities": [],
    })

    all_topics: set[str] = set()

    for r in deduped:
        conf = r.get("confidence", "Low")
        weight = CONFIDENCE_WEIGHTS.get(conf, 1)
        member_name = r.get("member_name", "Unknown")
        party = r.get("party") or ""
        member_type = r.get("member_type") or ""
        source_type = r.get("source_type") or ""
        activity_date = r.get("activity_date") or ""

        topics_raw = r.get("topics")
        if isinstance(topics_raw, str):
            try:
                topics = json.loads(topics_raw)
            except (json.JSONDecodeError, TypeError):
                topics = [topics_raw] if topics_raw else []
        elif isinstance(topics_raw, list):
            topics = topics_raw
        else:
            topics = []

        all_topics.update(topics)
        act = _activity_record(r)

        # Update cross-topic entry
        cm = cross_members[member_name]
        cm["member_name"] = member_name
        cm["party"] = party
        cm["member_type"] = member_type
        cm["score"] += weight
        cm["mentions"] += 1
        cm["topics_active"].update(topics)
        cm["activities"].append(act)

        for topic in topics:
            tm = topic_members[topic][member_name]
            tm["member_name"] = member_name
            tm["party"] = party
            tm["member_type"] = member_type
            tm["score"] += weight
            tm["mentions"] += 1
            tm["sources"][source_type] += 1
            tm["activities"].append(act)

            if activity_date > tm["latest_date"]:
                tm["latest_date"] = activity_date

    # Sort activities newest-first within each member
    for _, members in topic_members.items():
        for tm in members.values():
            tm["activities"].sort(key=lambda a: a["activity_date"], reverse=True)

    for cm in cross_members.values():
        cm["activities"].sort(key=lambda a: a["activity_date"], reverse=True)

    # Serialise topic results
    topics_out: dict[str, list[dict]] = {}
    for topic, members in topic_members.items():
        ranked = []
        for i, (_, m) in enumerate(
            sorted(members.items(), key=lambda x: (-x[1]["mentions"], x[0])), 1
        ):
            ranked.append({
                "rank": i,
                "member_name": m["member_name"],
                "party": m["party"],
                "member_type": m["member_type"],
                "score": m["score"],
                "mentions": m["mentions"],
                "sources": dict(m["sources"]),
                "latest_date": m["latest_date"],
                "activities": m["activities"],
            })
        topics_out[topic] = ranked

    # Serialise cross-topic results
    cross_out = []
    for i, (_, m) in enumerate(
        sorted(cross_members.items(), key=lambda x: (-x[1]["mentions"], x[0])), 1
    ):
        topics_active = sorted(m["topics_active"])
        cross_out.append({
            "rank": i,
            "member_name": m["member_name"],
            "party": m["party"],
            "member_type": m["member_type"],
            "topics_active": len(topics_active),
            "topic_list": topics_active,
            "score": m["score"],
            "mentions": m["mentions"],
            "activities": m["activities"],
        })

    return {
        "topics": topics_out,
        "cross_topic": cross_out,
        "meta": {
            "total_results": len(deduped),
            "total_members": len(cross_members),
            "topics": sorted(all_topics),
        },
    }


def create_index_excel(index_data: dict, scan_summaries: list[dict]) -> BytesIO:
    """Generate an Excel workbook: one sheet per topic + cross-topic summary.

    Requires openpyxl. Returns BytesIO buffer ready to send.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError("openpyxl is required for Excel export") from e

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1A365D")
    header_align = Alignment(horizontal="center", wrap_text=True)

    def _style_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def _autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    # Per-topic sheets
    for topic, members in index_data.get("topics", {}).items():
        sheet_name = topic[:31]
        ws = wb.create_sheet(title=sheet_name)
        _style_header(ws, [
            "Rank", "Member", "Party", "Type", "Score", "Mentions",
            "Sources", "Latest Activity",
        ])
        for m in members:
            sources_str = ", ".join(
                f"{SOURCE_LABELS.get(k, k)} ({v})"
                for k, v in sorted(m["sources"].items())
            )
            ws.append([
                m["rank"],
                m["member_name"],
                m["party"],
                m["member_type"],
                m["score"],
                m["mentions"],
                sources_str,
                m["latest_date"],
            ])
        _autofit(ws)

    # Cross-topic summary sheet
    ws_cross = wb.create_sheet(title="Cross-Topic Summary")
    _style_header(ws_cross, [
        "Rank", "Member", "Party", "Type",
        "Topics Active", "Topic List", "Total Score", "Total Mentions",
    ])
    for m in index_data.get("cross_topic", []):
        ws_cross.append([
            m["rank"],
            m["member_name"],
            m["party"],
            m["member_type"],
            m["topics_active"],
            ", ".join(m["topic_list"]),
            m["score"],
            m["mentions"],
        ])
    _autofit(ws_cross)

    # Scans included sheet
    ws_scans = wb.create_sheet(title="Scans Included")
    _style_header(ws_scans, ["Scan ID", "Start Date", "End Date", "Topics", "Results"])
    for s in scan_summaries:
        ws_scans.append([
            s.get("id"),
            s.get("start_date"),
            s.get("end_date"),
            ", ".join(s.get("topic_names", [])),
            s.get("result_count", 0),
        ])
    _autofit(ws_scans)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
